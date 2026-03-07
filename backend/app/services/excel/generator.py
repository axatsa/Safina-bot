import io
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_smeta_excel(data: dict) -> io.BytesIO:
    """
    Generates an Excel (.xlsx) SMETA document based on the provided data.
    `data` should be a dict containing:
    - request_id
    - date (str)
    - sender_name
    - sender_position
    - purpose
    - items (list of dicts: name, quantity, price, total)
    - total_amount
    - currency
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Смета"

    # Define Styles
    header_font = Font(name="Arial", size=14, bold=True)
    bold_font = Font(name="Arial", size=11, bold=True)
    regular_font = Font(name="Arial", size=11)
    
    center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_aligned = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_aligned = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Column Widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = f"СМЕТА РАСХОДОВ #{data.get('request_id')}"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_aligned
    
    row_idx = 3

    # Metadata
    meta_info = [
        ("Дата:", data.get('date')),
        ("Инициатор:", f"{data.get('sender_name')} ({data.get('sender_position')})"),
        ("Цель:", data.get('purpose')),
        ("Валюта:", data.get('currency')),
    ]

    for label, val in meta_info:
        ws.merge_cells(f'A{row_idx}:B{row_idx}')
        ws[f'A{row_idx}'] = label
        ws[f'A{row_idx}'].font = bold_font
        ws[f'A{row_idx}'].alignment = left_aligned
        
        ws.merge_cells(f'C{row_idx}:E{row_idx}')
        ws[f'C{row_idx}'] = val
        ws[f'C{row_idx}'].font = regular_font
        ws[f'C{row_idx}'].alignment = left_aligned
        row_idx += 1

    row_idx += 1 # Empty row
    
    # Table Headers
    headers = ["№", "Наименование", "Кол-во", "Цена за ед.", "Сумма"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = bold_font
        cell.alignment = center_aligned
        cell.border = thin_border
        cell.fill = header_fill
        
    row_idx += 1

    # Table Items
    items = data.get('items', [])
    for idx, item in enumerate(items, start=1):
        ws.cell(row=row_idx, column=1, value=idx).alignment = center_aligned
        ws.cell(row=row_idx, column=2, value=item.get('name', '')).alignment = left_aligned
        ws.cell(row=row_idx, column=3, value=float(item.get('quantity', 0))).alignment = center_aligned
        ws.cell(row=row_idx, column=4, value=float(item.get('price', 0))).alignment = right_aligned
        ws.cell(row=row_idx, column=5, value=float(item.get('total', 0))).alignment = right_aligned
        
        for col_idx in range(1, 6):
            ws.cell(row=row_idx, column=col_idx).border = thin_border
            ws.cell(row=row_idx, column=col_idx).font = regular_font
            
        row_idx += 1

    # Total Row
    ws.merge_cells(f'A{row_idx}:D{row_idx}')
    total_label_cell = ws.cell(row=row_idx, column=1, value="ИТОГО:")
    total_label_cell.font = bold_font
    total_label_cell.alignment = right_aligned
    
    total_value_cell = ws.cell(row=row_idx, column=5, value=float(data.get('total_amount', 0)))
    total_value_cell.font = bold_font
    total_value_cell.alignment = right_aligned
    
    for col_idx in range(1, 6):
        ws.cell(row=row_idx, column=col_idx).border = thin_border

    # Return as BytesIO
    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    
    return virtual_workbook
