import io
import datetime
import os
from decimal import Decimal
import pandas as pd
from sqlalchemy.orm import Session
from app.db import models
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

STATUS_MAP = {
    "request": "Запрос",
    "review": "На рассмотрении",
    "pending_senior": "На согласовании CFO",
    "approved_senior": "Утверждено CFO",
    "pending_ceo": "На согласовании CEO",
    "approved_ceo": "Одобрено CEO",
    "confirmed": "Подтверждено",
    "declined": "Отклонено",
    "revision": "Возврат на доработку",
    "archived": "Архивировано"
}

def generate_expenses_xlsx(expenses: list[models.ExpenseRequest]) -> io.BytesIO:
    data = []
    for e in expenses:
        usd_rate = Decimal(str(e.usd_rate)) if e.usd_rate else None
        items = e.items if isinstance(e.items, list) else []
        for item in items:
            item_currency = item.get("currency", e.currency)
            item_amount_native = Decimal(str(item.get("amount", 0))) * Decimal(str(item.get("quantity", 0)))
            
            amount_uzs = item_amount_native
            if item_currency == "USD" and usd_rate:
                amount_uzs = item_amount_native * usd_rate
            
            data.append({
                "ID Запроса": e.request_id,
                "Дата": e.date.strftime("%d.%m.%Y %H:%M"),
                "Проект": f"{e.project_name} ({e.project_code})" if e.project_name else "Без проекта",
                "Цель расхода": item.get("name") if len(items) > 1 else e.purpose,
                "Сумма": float(item_amount_native),
                "Валюта": item_currency,
                "Курс USD": float(usd_rate) if usd_rate else None,
                "Сумма в UZS": float(round(amount_uzs, 2)),
                "Ответственный": e.created_by,
                "Статус": STATUS_MAP.get(e.status, e.status)
            })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Expenses')
        worksheet = writer.sheets['Expenses']
        
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Styles for header
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align
            
        # Borders for content
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
        # Auto-adjust columns width
        for idx, col in enumerate(df.columns):
            if not df.empty:
                # Robustly calculate max string length in column
                series = df[col].dropna().astype(str)
                if not series.empty:
                    max_content_len = series.apply(len).max()
                else:
                    max_content_len = 0
                max_len = max(max_content_len, len(str(col))) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50)
            
        # Add SUM formulas if not empty
        if not df.empty:
            total_row = len(df) + 2
            worksheet.cell(row=total_row, column=7, value="ИТОГО:").font = bold_font
            worksheet.cell(row=total_row, column=7).border = thin_border
            
            # UZS Sum (col 8)
            sum_cell_uzs = worksheet.cell(row=total_row, column=8)
            sum_cell_uzs.value = f"=SUM(H2:H{total_row-1})"
            sum_cell_uzs.font = bold_font
            sum_cell_uzs.border = thin_border

    output.seek(0)
    return output

